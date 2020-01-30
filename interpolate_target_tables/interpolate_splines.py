# -*- coding: utf-8 -*-
"""
Created on Thu Mar 14 17:30:05 2019

@author: Tomer Fishman
"""

#@todo: sometimes the excel loading functions miss values that are excel functions. To sidestep this bug, open and save the excel file before running the script.

"""
# basic usage: 
# 0. ensure that a copy of cover.xlsx is in the same folder as the target tables excel file you wish to interpolate
# 1. in cell "#%% basic information" (row 34) edit the following: 
#   a. working directory (where the target tables excel file is located)
#   b. target tables excel file name
#   c. your name
# 2. in cell "#%% ok go!" around row 206 choose which sheets to process (use the comments just above it for examples)
# 3. run the entire script!
# 4. wait
# 5. The intperpolated & smoothed files will be in a new directory with the excel file's name
# 6. Check the log.xlsx to see what worked and what not
# 7. Contact tomer tomer.fishman@idc.ac.il or tomerfishman@gmail.com if something didn't work
"""

#%% load packages & stuff
import numpy as np
import pandas as pd
from scipy.interpolate import make_interp_spline
import os
import uuid
from openpyxl import load_workbook
#import matplotlib.pyplot as plt

#%% basic information:
workingdir = 'C:\\Users\\Tomer Fishman\\Dropbox (Personal)\\-the research\\2018 1 unep irp g7 project\\sandbox interpolate and smooth'
#workingdir = 'C:\\Users\\Tomer Fishman\\Dropbox (Personal)\\-the research\\2018 1 unep irp g7 project\\G7 RECC\\scenarios\\'
#workingdir = 'C:\\Users\\spauliuk.AD\\Dropbox\\G7 RECC\scenarios\\'

yourname = "Tomer Fishman" #will appear in the interpolated files' cover sheet
#yourname = "Stefan Pauliuk" #will appear in the interpolated files' cover sheet

excelfilename = "scenario_target_tables_MASTER.xlsx"

#%% functions

#debug log function
def logger(message):
    global log # use the global log variable
    print(message)
    log.append(message)

#interpolate and smooth a single table
def interpsmooth(sheetname, sheetdata, colLoc, rowLoc): #expand targets to a full time series ###
    try:
        product_g = sheetdata.iloc[0,colLoc] #get the value of g
        tableset = sheetdata.iloc[rowLoc:rowLoc+39,colLoc:colLoc+9] #work only with one table set
        region = tableset.iloc[0,1] #get the value of r
        logger("--Table in column "+str(colLoc)+" row "+str(rowLoc)+" ("+str(product_g)+", "+str(region)+")")
    
        x = np.array([2015,2020,2030,2040,2050,2060,2100]) #all potential years
        year = np.arange(2015,2101) #full time series, without 2014
        
        yLED = np.full(7, np.nan) #create empty arrays for the three scenarios
        ySSP1 = np.full(7, np.nan) #repeat for ssp1 and 2 below
        ySSP2 = np.full(7, np.nan)
        
        #check if compulsory values are missing in the target table
        nanYears = ''
        if str(tableset.iloc[5,3]) == 'nan': 
            nanYears += '2015 '
        if str(tableset.iloc[19,3]) == 'nan': 
            nanYears += '2060LED '
        if str(tableset.iloc[28,3]) == 'nan': 
            nanYears += '2060SSP1 '
        if str(tableset.iloc[37,3]) == 'nan': 
            nanYears += '2060SSP2 '
    
        if nanYears == '': #proceed only if all compulsory values are there
   
            yLED[0] = ySSP1[0] = ySSP2[0] = tableset.iloc[5,3] #historical 2015 for the three scenarios
            yLED[1:] = tableset.iloc[15:21,3] #2020 - 2050
            ySSP1[1:] = tableset.iloc[24:30,3] #repeat for ssp1 and 2 below
            ySSP2[1:] = tableset.iloc[33:39,3]
                
            xLED = x[~np.isnan(yLED)] #only the years whose values are not nan (the symbol ~ inverts the boolean answer)
            yLED = yLED[~np.isnan(yLED)] #only the values that are not nan (the symbol ~ inverts the boolean answer)
            xSSP1 = x[~np.isnan(ySSP1)] #repeat for ssp1 and 2 below
            ySSP1 = ySSP1[~np.isnan(ySSP1)]
            xSSP2 = x[~np.isnan(ySSP2)]
            ySSP2 = ySSP2[~np.isnan(ySSP2)]

            pos_final = np.where(yLED == yLED[-1])[0][0] #determine final target value
            timeseriesLED = np.full(year.size, yLED[pos_final]) #create full series padded with the final target value
            if pos_final != 0: #proceed only if the target value is not the 2015 value (i.e. if there's change over time)
                pos_year = int(np.where(year == xLED[pos_final])[0]) #determine final target's year's position
                clamped_spline = make_interp_spline(xLED[:pos_final+1],yLED[:pos_final+1], bc_type=([(2, 0)], [(1, 0)])) #spline function, free (2nd derivative=0) for starting boundary condition and clamped (1st derivative=0) for end boundary condition
                timeseriesLED[:pos_year+1] = clamped_spline(year[:pos_year+1]) #fill the full series with the interpolated spline values

            pos_final = np.where(ySSP1 == ySSP1[-1])[0][0] #repeat for ssp1
            timeseriesSSP1 = np.full(year.size, ySSP1[pos_final])
            if pos_final != 0:
                pos_year = int(np.where(year == xSSP1[pos_final])[0])
                clamped_spline = make_interp_spline(xSSP1[:pos_final+1],ySSP1[:pos_final+1], bc_type=([(2, 0)], [(1, 0)]))
                timeseriesSSP1[:pos_year+1] = clamped_spline(year[:pos_year+1])

            pos_final = np.where(ySSP2 == ySSP2[-1])[0][0] #repeat for SSP2
            timeseriesSSP2 = np.full(year.size, ySSP2[pos_final])
            if pos_final != 0:
                pos_year = int(np.where(year == xSSP2[pos_final])[0])
                clamped_spline = make_interp_spline(xSSP2[:pos_final+1],ySSP2[:pos_final+1], bc_type=([(2, 0)], [(1, 0)]))
                timeseriesSSP2[:pos_year+1] = clamped_spline(year[:pos_year+1])

        
            combined = pd.DataFrame({'g' : [product_g] * len(year), 'R' : [region] * len(year), 't' : year, 'LED' : timeseriesLED, 'SSP1' : timeseriesSSP1, 'SSP2' : timeseriesSSP2}) #format a nice pandas dataframe
            logger("  Interpolated!")
    
        else: #report if this table could not be processed because values are missing
            combined = pd.DataFrame()
            logger("  "+nanYears+"empty so skipped")
    except: #report if this table doesn't fit the expected format
        combined = pd.DataFrame()
        logger("  erronous format or missing g or r info")
    return combined

# iterate through all sheets of an Excel file
def exceliterator(excelfile, sheetNames):
    wb = load_workbook(excelfilename)
    interp_date = pd.to_datetime('today').strftime('%Y-%m-%d')
    
    for name in sheetNames: #cycle through all sheets in the excel file

        logger("<"+name+">") #log the name of the sheet
        excelsheet = excelfile[name]
        
        if excelsheet.iloc[0,0] == 'Blue cells are optional.': # and excelsheet.iloc[0,3] == 'Interpolation date': #check if the sheet is a target table sheet
            tableset_cornersCols = np.arange(6,excelsheet.shape[1],12) # column number of the left-hand edges of all tables
            tableset_cornersRows = np.arange(2,excelsheet.shape[0],41) # row number of the top edges of all tables
            logmessage = "--Detected "+str(len(tableset_cornersCols) * len(tableset_cornersRows))+" tables in columns "+str(tableset_cornersCols)+" rows "+str(tableset_cornersRows)
            logger(logmessage)
            interpolatedlong = pd.DataFrame() #start with an empty dataframe
            interpolatedlong = pd.concat([interpsmooth(name,excelsheet,i,j) for i in tableset_cornersCols for j in tableset_cornersRows],ignore_index=True) #loop through all tablesets in the sheet using nested list comprehensions.
            if not(interpolatedlong.empty): #because interpsmooth returns empty dataframes if no tables were interpolated.
                
                #create Cover sheet:
                coversheet = excelcover['Cover']
                interp_uuid = str(uuid.uuid4())
                coversheet['B3'] = name #Dataset_Name
                coversheet['B4'] = str(excelsheet.iloc[3,1])+". "+str(excelsheet.iloc[2,1])+". "+str(excelsheet.iloc[4,1])+". "+str(excelsheet.iloc[6,1])+". "+str(excelsheet.iloc[5,1])+". (this info extracted from the target table excel sheet, refer to it for details.)" #Dataset_Description
                coversheet['B5'] = excelsheet.iloc[6,10] #Dataset_Unit
                coversheet['B7'] = 'Interpolated and MA-smoothed from target table file "'+excelfilename+'". Please refer to it for details.' #Dataset_Comment
                coversheet['B10'] = interp_uuid #Dataset_UUID
                coversheet['B12'] = interp_date #Last modified
                coversheet['B13'] = yourname #Last modified by
                coversheet['B23'] = excelsheet.iloc[8,0] #Row Aspects_Meaning
                coversheet['D21'] = len(interpolatedlong.index) #Col Aspects_Meaning
                coversheet['F24'] = excelsheet.iloc[7,10] #DATA_Info unit
                excelcover.save(exportfolder+name+'.xlsx')
               
                #create values sheet
                writer = pd.ExcelWriter(exportfolder+name+'.xlsx', engine='openpyxl', mode='a')
                interpolatedlong.to_excel(writer, index = False, sheet_name = "values")
                
                #save a copy of the target table sheet to the newly created excel file
                #excelsheet.to_excel(writer, index = False, sheet_name='target tables')
                writer.save()
                
                #add this run's date & uuid to the target table excel file
                lastvalidrow = str(excelsheet['Interpolation history'].last_valid_index()+3)
                ws = wb[name]
                ws['C'+lastvalidrow] = interp_uuid
                ws['D'+lastvalidrow] = interp_date
                
                #report back
                logmessage = "Exported with UUID "+interp_uuid
            else: logmessage = "Returned empty and not exported"
        else: logmessage = "Is not a target table template sheet"
        logger(logmessage)
        logger("")
    #save all runs' date & uuid to the target table excel file
    wb.save(excelfilename)

#%% ok go!

#set up working files and directories
os.chdir(workingdir)

#initialize log
log = list()
logger("Started "+str(pd.to_datetime('today')))

#set up the export folder
exportfolder = "interpolated_"+excelfilename[0:-5]+"\\"
if not os.path.exists(exportfolder):
    os.makedirs(exportfolder)
    #os.mkdir(exportfolder+"figures\\")
    logmessage = "Created directory "+exportfolder
else: logmessage = "Saving to existing directory "+exportfolder+" overwriting any previous files with identical names already there"
logger(logmessage)
logger("")

#load excel files (target tables and cover template)
excelimport = pd.read_excel(excelfilename, None)
excelcover = load_workbook("cover.xlsx")

#get sheet names
sheets = list(excelimport)

#this is your chance to remove sheets from processing
#for instance: 
#sheets[:] will process all sheets
#sheets[23:24] for a single sheet (typing only sheets[23] returns a string, not a list. sheets[23:23] returns an empty list)
#sheets[23:30] will process a series of sheets
processsheets = sheets[:37]

#extrapolate and save export files
exceliterator(excelimport, processsheets)

#finalize and save log
logger("Completed "+str(pd.to_datetime('today')))
writer_log = pd.ExcelWriter(exportfolder+'0log.xlsx')
pd.Series(log).to_excel(writer_log, index = False, header = False)
writer_log.save()
