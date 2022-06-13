# -*- coding: utf-8 -*-
"""
Created on Tue Jun  9 19:34:52 2020

@author: spauliuk
"""

import os
import openpyxl

mywb  = openpyxl.load_workbook(os.path.join('RECC_scenario_target_tables_v2_4.xlsx'))

###################################################################

# insert equations into 3_SHA_TypeSplit_Buildings

#Ssheet = mywb['3_SHA_TypeSplit_Buildings']
#
#TargetSheetCols = ['J','V','AH','AT','BF','BR','CD','CP','DB','DN','DZ','EL','EX']
#TargetSheetRows = [9,22,31,40,50,63,72,81,91,104,113,122,132,145,154,163,173,186,195,204,214,227,236,245,255,268,277,286,296,309,318,327, \
#                   337,350,359,368,378,391,400,409,419,432,441,450,460,473,482,491,501,514,523,532,\
#                   542,555,564,573,583,596,605,614,624,637,646,655,665,678,687,696,706,719,728,737,747,760,769,778,788,801,810,819]         
#
#SourceSheetCols = ['D','E','F','G','H','I','J','K','L','M','N','O','P']
#SourceSheetRows = ['30','32','33','34','45','47','48','49','60','62','63','64','75','77','78','79','90','92','93','94','105','107','108','109','120','122','123','124','135','137','138','139','150','152','153','154',\
#                   '165','167','168','169','180','182','183','184','195','197','198','199','210','212','213','214',\
#                   '225','227','228','229','270','272','273','274','285','287','288','289','240','242','243','244','300','302','303','304','255','257','258','259','315','317','318','319']         
#
#for m in range(0,len(TargetSheetRows)):
#    for n in range(0,len(TargetSheetCols)):
#        Ssheet[TargetSheetCols[n]+str(TargetSheetRows[m])]    = '=BackgrndCalc_Type_Split_Bld!' + SourceSheetCols[n] + SourceSheetRows[m] # 2050 value
#        Ssheet[TargetSheetCols[n]+str(TargetSheetRows[m]+1)]  = '=BackgrndCalc_Type_Split_Bld!' + SourceSheetCols[n] + SourceSheetRows[m] # 2060 value
     

###################################################################

# insert equations into 3_SHA_TypeSplit_Buildings
        
Ssheet = mywb['3_SHA_DownSizing_Vehicles']

TargetSheetCols = ['J','V','AH','AT'] # 4 lateral entries
TargetSheetRows = [i + 41*n for n in range(0,21) for i in [20,29,38]] # [20,29,38,61,70,79,102,111,120,...] # 2030 cells for all 21 regions.

for m in range(0,len(TargetSheetRows)):
    for n in range(0,len(TargetSheetCols)):
        Ssheet[TargetSheetCols[n]+str(TargetSheetRows[m])]    = '' # 2030 value: empty
        Ssheet[TargetSheetCols[n]+str(TargetSheetRows[m]+1)]  = TargetSheetCols[n]+str(TargetSheetRows[m]+2)# 2040 value = 2050 value   

mywb.save('RECC_scenario_target_tables_v2_4_NEW_MASTER.xlsx')

# Test:
#Ssheet.cell(row = 1,  column = 3).value  = '=Tabelle2!C3'
#Ssheet.cell(row = 2,  column = 3).value  = '=Tabelle2!C3*Tabelle2!C4'
#Ssheet.cell(row = 3,  column = 3).value  = '=Tabelle2!C3+Tabelle2!C5'
#Ssheet.cell(row = 4,  column = 3).value  = '=Tabelle2!C4+Tabelle2!C5*Tabelle2!C6'